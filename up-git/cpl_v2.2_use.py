# -*- coding: utf-8 -*-
"""
Excel日期比对工具
比较Excel文件中N列和AC列的日期，将年月不同的行复制到新的sheet
如果AC列日期比N列日期更靠后，则将该行标记为红色背景
在第31列前插入「確認結果」列，保留原样式，设置数据验证
"""

import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
from copy import copy
import sys
import os

# 设置默认编码
if sys.version_info[0] < 3:
    reload(sys)
    sys.setdefaultencoding('utf-8')

print("GAME START")

def extract_year_month(val):
    if isinstance(val, datetime):
        return val.strftime("%Y-%m")
    if isinstance(val, str) and val:
        val = val.strip()
        try:
            for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y-%m", "%Y/%m"]:
                try:
                    return datetime.strptime(val, fmt).strftime("%Y-%m")
                except:
                    pass
        except:
            pass
    return None

def parse_date_for_comparison(val):
    if isinstance(val, datetime):
        return val
    if isinstance(val, str) and val:
        val = val.strip()
        try:
            for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y-%m", "%Y/%m"]:
                try:
                    return datetime.strptime(val, fmt)
                except:
                    pass
        except:
            pass
    return None

def copy_cell(source, target):
    target.value = source.value
    if source.has_style:
        target.font = copy(source.font)
        target.border = copy(source.border)
        target.fill = copy(source.fill)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)

def apply_red_background(ws, row_num, max_col):
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    for col in range(1, max_col + 1):
        ws.cell(row_num, col).fill = red_fill

def main():
    print("Main function started")

    root = tk.Tk()
    root.withdraw()

    file_path = filedialog.askopenfilename(
        title="Excelファイルを選択してください",
        filetypes=[("Excelファイル", "*.xlsx *.xlsm *.xls")]
    )

    if not file_path:
        print("ファイルが選択されていません")
        return

    print(f"選択されたファイル: {file_path}")

    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
        sheet_names = wb.sheetnames

        if len(sheet_names) == 1:
            sheet_name = sheet_names[0]
        else:
            select_window = tk.Toplevel()
            select_window.title("シートを選択")
            select_window.geometry("300x400")

            tk.Label(select_window, text="処理対象のシートを選択してください:").pack(pady=10)

            listbox = tk.Listbox(select_window, height=15)
            listbox.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

            for name in sheet_names:
                listbox.insert(tk.END, name)

            listbox.selection_set(0)
            selected_sheet = {"name": None}

            def on_select():
                selection = listbox.curselection()
                if selection:
                    selected_sheet["name"] = sheet_names[selection[0]]
                    select_window.destroy()

            def on_double_click(event):
                on_select()

            listbox.bind("<Double-Button-1>", on_double_click)
            tk.Button(select_window, text="OK", command=on_select).pack(pady=10)

            select_window.wait_window()
            sheet_name = selected_sheet["name"]

            if not sheet_name:
                print("シートが選択されていません")
                return

        print(f"選択されたシート:  {sheet_name}")
        ws = wb[sheet_name]

        # 删除已存在的 Different_dates sheet
        if 'Different_dates' in wb.sheetnames:
            del wb['Different_dates']
        new_ws = wb.create_sheet('Different_dates')

        N_COL = 14  # Column N
        AC_COL = 29  # Column AC
        original_max_col = ws.max_column

        # 首先完全复制原始表格结构（前两行：第1行空行，第2行表头）
        for row in range(1, 3):
            for col in range(1, original_max_col + 1):
                copy_cell(ws.cell(row, col), new_ws.cell(row, col))

        # 收集符合条件的数据行并复制
        new_row = 3  # 从第3行开始（保持表头结构）
        count = 0
        red_count = 0

        for orig_row in range(3, ws.max_row + 1):
            n_val = ws.cell(orig_row, N_COL).value
            ac_val = ws.cell(orig_row, AC_COL).value
            
            n_date = extract_year_month(n_val)
            ac_date = extract_year_month(ac_val)
            
            if n_date and ac_date and n_date != ac_date:
                count += 1
                
                # 复制整行数据
                for col in range(1, original_max_col + 1):
                    copy_cell(ws.cell(orig_row, col), new_ws.cell(new_row, col))
                
                # 检查是否需要标记为红色
                n_datetime = parse_date_for_comparison(n_val)
                ac_datetime = parse_date_for_comparison(ac_val)
                if n_datetime and ac_datetime and ac_datetime > n_datetime:
                    apply_red_background(new_ws, new_row, original_max_col)
                    red_count += 1
                
                new_row += 1

        if count == 0:
            messagebox.showinfo("Complete", "年月が異なる行は見つかりませんでした。")
            return

        # 现在在所有数据复制完成后插入第31列
        new_ws.insert_cols(31)

        # 设置第31列的表头
        # 第1行第31列保持空
        copy_cell(new_ws.cell(1, 30), new_ws.cell(1, 31))
        new_ws.cell(1, 31).value = None
        
        # 第2行第31列设置为"確認結果"
        copy_cell(new_ws.cell(2, 30), new_ws.cell(2, 31))
        new_ws.cell(2, 31).value = "確認結果"

        # 为所有数据行的第31列设置样式和空值
        for row_idx in range(3, new_row):
            copy_cell(new_ws.cell(row_idx, 30), new_ws.cell(row_idx, 31))
            new_ws.cell(row_idx, 31).value = None

        # 复制列宽
        for col_letter in ws.column_dimensions:
            col_num = openpyxl.utils.column_index_from_string(col_letter)
            if col_num <= 30:
                new_ws.column_dimensions[col_letter].width = ws.column_dimensions[col_letter].width
            elif col_num >= 31:
                # 原来的31列及之后现在变成32列及之后
                new_col_letter = openpyxl.utils.get_column_letter(col_num + 1)
                new_ws.column_dimensions[new_col_letter].width = ws.column_dimensions[col_letter].width
        
        # 设置新的確認結果列（31列）的宽度
        new_ws.column_dimensions['AE'].width = 12

        # 复制合并单元格（仅表头部分）
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row <= 2:
                # 如果合并区域包含31列及之后，需要调整
                if merged_range.min_col >= 31:
                    # 整个合并区域在31列之后，所有列号+1
                    new_ws.merge_cells(
                        start_row=merged_range.min_row, start_column=merged_range.min_col + 1,
                        end_row=merged_range.max_row, end_column=merged_range.max_col + 1
                    )
                elif merged_range.max_col >= 31:
                    # 合并区域跨越31列，需要分割处理
                    if merged_range.min_col < 31:
                        # 31列之前的部分
                        new_ws.merge_cells(
                            start_row=merged_range.min_row, start_column=merged_range.min_col,
                            end_row=merged_range.max_row, end_column=30
                        )
                        # 32列之后的部分
                        if merged_range.max_col > 31:
                            new_ws.merge_cells(
                                start_row=merged_range.min_row, start_column=32,
                                end_row=merged_range.max_row, end_column=merged_range.max_col + 1
                            )
                else:
                    # 合并区域在31列之前，直接复制
                    new_ws.merge_cells(str(merged_range))

        # 创建数据验证对象（尝试更直接的设置方式）
        from openpyxl.worksheet.datavalidation import DataValidation
        
        # 使用更明确的参数设置
        dv = DataValidation(
            type="list",
            formula1='"計画納期が正,出荷予定日が正,工程調査"',
            allow_blank=True
        )
        
        # 手动设置showDropDown属性
        dv.showDropDown = True
        dv.showInputMessage = False
        dv.showErrorMessage = False
        
        # 创建单元格范围字符串
        if new_row > 3:
            range_str = f"AE3:AE{new_row-1}"
            dv.add(range_str)
            new_ws.add_data_validation(dv)
            
            # 额外尝试：为每个单元格单独设置
            for row_idx in range(3, new_row):
                cell = new_ws.cell(row_idx, 31)
                # 尝试直接设置单元格的数据验证属性
                try:
                    cell.data_type = 's'  # 设置为字符串类型
                except:
                    pass

        wb.save(file_path)
        wb.close()

        message = f"{count} 件の年月が異なるデータが見つかりました。\n"
        message += f"そのうち, {red_count} 件は、出荷予定日が計画納期より遅く、赤色でマークされました。\n"
        message += "列31（AE）に「確認結果」列を追加し、選択肢付きの入力制限を設定しました。\n"
        message += "結果は「Different_dates」シートに保存されました。"

        messagebox.showinfo("Complete", message)

    except Exception as e:
        import traceback
        error_msg = f"Error: {str(e)}\n\nDetails:\n{traceback.format_exc()}"
        print(error_msg)
        messagebox.showerror("Error", error_msg)

if __name__ == "__main__":
    print("Excel Date Comparison Tool")
    try:
        main()
    except Exception as e:
        print(f"Error: {e}")
        input("Press Enter to exit...")
    print("GAMEOVER")