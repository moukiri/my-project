import cv2
import numpy as np
import fitz
import pytesseract
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import re
import logging
from pathlib import Path
from typing import List, Dict, Tuple, Optional
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import sys
import os

# 配置 Tesseract 路径
def configure_tesseract():
    """配置 Tesseract OCR 路径"""
    possible_paths = [
        r'C:\Program Files\Tesseract-OCR\tesseract.exe',
        r'C:\Program Files (x86)\Tesseract-OCR\tesseract.exe',
        r'C:\Users\{}\AppData\Local\Tesseract-OCR\tesseract.exe'.format(os.getenv('USERNAME', '')),
        '/usr/bin/tesseract',
        '/usr/local/bin/tesseract',
        '/opt/homebrew/bin/tesseract'
    ]
    
    # 首先检查是否已经配置了路径
    try:
        pytesseract.get_tesseract_version()
        print("✅ Tesseract OCR 已正确配置")
        return True
    except:
        pass
    
    # 尝试常见路径
    for path in possible_paths:
        if os.path.exists(path):
            pytesseract.pytesseract.tesseract_cmd = path
            try:
                pytesseract.get_tesseract_version()
                print(f"✅ 找到 Tesseract OCR: {path}")
                return True
            except:
                continue
    
    print("❌ 无法自动找到 Tesseract OCR")
    return False

# 在程序开始时配置 Tesseract
if not configure_tesseract():
    root = tk.Tk()
    root.withdraw()
    
    messagebox.showinfo("配置 Tesseract OCR", 
                       "请选择 tesseract.exe 文件\n"
                       "通常位于: C:\\Program Files\\Tesseract-OCR\\tesseract.exe")
    
    tesseract_path = filedialog.askopenfilename(
        title="选择 tesseract.exe 文件",
        filetypes=[("可执行文件", "*.exe"), ("所有文件", "*.*")]
    )
    
    if tesseract_path and os.path.exists(tesseract_path):
        pytesseract.pytesseract.tesseract_cmd = tesseract_path
        try:
            pytesseract.get_tesseract_version()
            print(f"✅ 手动配置成功: {tesseract_path}")
        except:
            print("❌ 选择的文件不是有效的 Tesseract OCR")
            sys.exit(1)
    else:
        print("❌ 未选择有效的 Tesseract OCR 文件")
        sys.exit(1)
    
    root.destroy()

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('pdf_recognition.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class MonthMarkingRecognizer:
    def __init__(self):
        self.valid_months = ['1月', '2月', '3月', '4月', '5月', '6月', 
                            '7月', '8月', '9月', '10月', '11月', '12月']
        self.debug_enabled = True  # 启用调试模式
    
    def save_debug_image(self, image, filename):
        """保存调试图片"""
        if self.debug_enabled:
            try:
                debug_filename = f"debug_{filename}"
                cv2.imwrite(debug_filename, image)
                logger.info(f"保存调试图片: {debug_filename}")
            except Exception as e:
                logger.warning(f"保存调试图片失败: {e}")
    
    def detect_red_markings_improved(self, image: np.ndarray) -> List[Tuple[int, int, int, int]]:
        """改进的红色标记检测 - 专门针对表格中的红色圆圈"""
        logger.info("开始改进的红色标记检测...")
        
        # 转换到HSV颜色空间
        hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
        
        # 更严格的红色范围 - 针对鲜艳的红色
        lower_red1 = np.array([0, 150, 150])    # 更高的饱和度和亮度要求
        upper_red1 = np.array([10, 255, 255])
        lower_red2 = np.array([170, 150, 150])  # 更严格的红色范围
        upper_red2 = np.array([180, 255, 255])
        
        # 创建红色掩码
        mask1 = cv2.inRange(hsv, lower_red1, upper_red1)
        mask2 = cv2.inRange(hsv, lower_red2, upper_red2)
        red_mask = cv2.bitwise_or(mask1, mask2)
        
        # 保存原始掩码
        self.save_debug_image(red_mask, "red_mask_strict.png")
        
        # 形态学操作 - 去除小噪点
        kernel_small = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))
        red_mask = cv2.morphologyEx(red_mask, cv2.MORPH_OPEN, kernel_small)
        
        # 闭合操作 - 连接断开的部分
        kernel_large = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5, 5))
        red_mask = cv2.morphologyEx(red_mask, cv2.MORPH_CLOSE, kernel_large)
        
        self.save_debug_image(red_mask, "red_mask_processed.png")
        
        # 寻找轮廓
        contours, _ = cv2.findContours(red_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        markings = []
        debug_image = image.copy()
        
        for i, contour in enumerate(contours):
            area = cv2.contourArea(contour)
            
            # 更严格的面积筛选 - 真正的红色圆圈应该有一定大小
            if 100 < area < 2000:  # 缩小面积范围，过滤掉太小的噪点
                # 计算轮廓属性
                perimeter = cv2.arcLength(contour, True)
                if perimeter > 0:
                    circularity = 4 * np.pi * area / (perimeter * perimeter)
                    
                    # 更严格的圆形度要求
                    if circularity > 0.5:  # 提高圆形度要求
                        x, y, w, h = cv2.boundingRect(contour)
                        aspect_ratio = w / h if h > 0 else 0
                        
                        # 更严格的长宽比要求 - 接近正圆
                        if 0.6 < aspect_ratio < 1.7:  # 更接近1:1的比例
                            # 检查周围是否有表格线条 - 这是关键改进
                            if self.is_near_table_structure(image, (x, y, w, h)):
                                markings.append((x, y, w, h))
                                # 标记找到的圆圈
                                cv2.rectangle(debug_image, (x, y), (x+w, y+h), (0, 255, 0), 3)
                                cv2.putText(debug_image, f"M{len(markings)}", (x, y-10), 
                                          cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)
                                logger.info(f"找到有效标记 {len(markings)}: 位置=({x},{y}), 大小=({w}x{h}), 面积={area:.1f}, 圆形度={circularity:.3f}")
        
        self.save_debug_image(debug_image, "markings_detected.png")
        logger.info(f"检测到 {len(markings)} 个有效的红色标记")
        return markings
    
    def is_near_table_structure(self, image: np.ndarray, region: Tuple[int, int, int, int]) -> bool:
        """检查标记是否靠近表格结构"""
        x, y, w, h = region
        
        # 扩大检查区域
        padding = 100
        x_start = max(0, x - padding)
        y_start = max(0, y - padding)
        x_end = min(image.shape[1], x + w + padding)
        y_end = min(image.shape[0], y + h + padding)
        
        roi = image[y_start:y_end, x_start:x_end]
        gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
        
        # 检测直线（表格线）
        edges = cv2.Canny(gray, 50, 150)
        lines = cv2.HoughLinesP(edges, 1, np.pi/180, threshold=50, minLineLength=50, maxLineGap=10)
        
        # 如果检测到足够的直线，认为是表格区域
        return lines is not None and len(lines) > 5
    
    def extract_table_content_improved(self, image: np.ndarray, markings: List[Tuple[int, int, int, int]]) -> List[Dict]:
        """改进的表格内容提取"""
        if not markings:
            return []
        
        logger.info("开始改进的表格内容提取...")
        
        # 全页OCR提取所有文本
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        
        # 提高图像质量
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
        gray = clahe.apply(gray)
        
        # OCR配置 - 针对日文和数字优化
        config = r'--oem 3 --psm 6 -c preserve_interword_spaces=1'
        
        try:
            # 获取OCR数据
            data = pytesseract.image_to_data(gray, lang='jpn+eng', config=config, output_type=pytesseract.Output.DICT)
            
            # 过滤低置信度文本
            filtered_data = {
                'text': [],
                'left': [],
                'top': [],
                'width': [],
                'height': [],
                'conf': []
            }
            
            for i, text in enumerate(data['text']):
                if int(data['conf'][i]) > 20 and text.strip():  # 降低置信度阈值
                    filtered_data['text'].append(text.strip())
                    filtered_data['left'].append(data['left'][i])
                    filtered_data['top'].append(data['top'][i])
                    filtered_data['width'].append(data['width'][i])
                    filtered_data['height'].append(data['height'][i])
                    filtered_data['conf'].append(data['conf'][i])
            
            logger.info(f"OCR提取到 {len(filtered_data['text'])} 个有效文本块")
            
            # 为每个标记查找相关的表格内容
            table_items = []
            
            for mark_idx, mark in enumerate(markings):
                mark_x, mark_y, mark_w, mark_h = mark
                mark_center_x = mark_x + mark_w // 2
                mark_center_y = mark_y + mark_h // 2
                
                logger.info(f"分析标记 {mark_idx + 1}: 中心位置 ({mark_center_x}, {mark_center_y})")
                
                # 查找标记左侧的文本（注番和项番通常在左侧）
                nearby_texts = []
                
                for i, text in enumerate(filtered_data['text']):
                    text_x = filtered_data['left'][i]
                    text_y = filtered_data['top'][i]
                    text_w = filtered_data['width'][i]
                    text_h = filtered_data['height'][i]
                    text_center_x = text_x + text_w // 2
                    text_center_y = text_y + text_h // 2
                    
                    # 检查文本是否在标记的左侧合理范围内
                    horizontal_distance = mark_center_x - text_center_x
                    vertical_distance = abs(mark_center_y - text_center_y)
                    
                    # 文本在标记左侧，且垂直距离不太远
                    if 10 < horizontal_distance < 300 and vertical_distance < 100:
                        nearby_texts.append({
                            'text': text,
                            'x': text_x,
                            'y': text_y,
                            'center_x': text_center_x,
                            'center_y': text_center_y,
                            'horizontal_dist': horizontal_distance,
                            'vertical_dist': vertical_distance,
                            'conf': filtered_data['conf'][i]
                        })
                
                # 按水平距离排序，找到最接近的文本
                nearby_texts.sort(key=lambda x: x['horizontal_dist'])
                
                if nearby_texts:
                    # 尝试识别注番和项番
                    note_number = None
                    item_number = None
                    
                    for text_info in nearby_texts[:5]:  # 只检查最近的5个文本
                        text = text_info['text']
                        
                        # 注番模式：T开头+6位数字
                        if re.match(r'^T\d{6}$', text):
                            note_number = text
                            logger.info(f"  找到注番: {text}")
                        
                        # 项番模式：JS开头+4位数字
                        elif re.match(r'^JS\s*\d{4}$', text.replace(' ', '')):
                            item_number = text.replace(' ', '')
                            logger.info(f"  找到项番: {text}")
                        
                        # 数字模式（可能是项番的一部分）
                        elif re.match(r'^\d{4}$', text):
                            # 查找附近的"JS"
                            for other_text in nearby_texts:
                                if other_text['text'] == 'JS' and abs(other_text['center_y'] - text_info['center_y']) < 20:
                                    item_number = f"JS{text}"
                                    logger.info(f"  组合项番: JS{text}")
                                    break
                    
                    # 如果找到了内容，添加到结果中
                    if note_number or item_number:
                        table_items.append({
                            'note_number': note_number or '未识别',
                            'item_number': item_number or '未识别',
                            'mark_position': mark,
                            'mark_index': mark_idx,
                            'nearby_texts': [t['text'] for t in nearby_texts[:3]]  # 保存附近文本用于调试
                        })
                        logger.info(f"  关联成功: 注番={note_number}, 项番={item_number}")
                    else:
                        logger.warning(f"  标记 {mark_idx + 1} 附近未找到有效的注番或项番")
                        logger.info(f"  附近文本: {[t['text'] for t in nearby_texts[:5]]}")
                else:
                    logger.warning(f"  标记 {mark_idx + 1} 附近无文本")
            
            logger.info(f"表格内容提取完成，找到 {len(table_items)} 个有效项目")
            return table_items
            
        except Exception as e:
            logger.error(f"表格内容提取失败: {e}")
            return []
    
    def extract_month_from_marking(self, image: np.ndarray, region: Tuple[int, int, int, int]) -> Optional[str]:
        """从红色标记中提取月份信息"""
        x, y, w, h = region
        
        # 扩大搜索区域 - 月份可能在标记内部或附近
        padding = 20
        x_start = max(0, x - padding)
        y_start = max(0, y - padding)
        x_end = min(image.shape[1], x + w + padding)
        y_end = min(image.shape[0], y + h + padding)
        
        roi = image[y_start:y_end, x_start:x_end]
        
        if roi.size == 0:
            return None
        
        # 保存原始ROI
        self.save_debug_image(roi, f"month_roi_{x}_{y}.png")
        
        # 转换为灰度
        gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
        
        # 多种预处理方法
        methods = [
            # 反转图像（白字黑底）
            lambda img: cv2.bitwise_not(img),
            # OTSU阈值
            lambda img: cv2.threshold(img, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1],
            # 反转OTSU
            lambda img: cv2.bitwise_not(cv2.threshold(img, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]),
            # 高阈值
            lambda img: cv2.threshold(img, 180, 255, cv2.THRESH_BINARY)[1],
            # 低阈值
            lambda img: cv2.threshold(img, 100, 255, cv2.THRESH_BINARY)[1]
        ]
        
        best_result = None
        best_confidence = 0
        
        for i, method in enumerate(methods):
            try:
                processed = method(gray)
                
                # 放大图像提高OCR精度
                scale = 6
                processed = cv2.resize(processed, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
                
                # 保存处理后的图像
                self.save_debug_image(processed, f"month_processed_{x}_{y}_method{i+1}.png")
                
                # OCR配置 - 针对数字和月字
                configs = [
                    r'--oem 3 --psm 8 -c tessedit_char_whitelist=0123456789月',  # 单字符
                    r'--oem 3 --psm 10 -c tessedit_char_whitelist=0123456789月', # 单字符无OSD
                    r'--oem 3 --psm 7 -c tessedit_char_whitelist=0123456789月',  # 单行
                ]
                
                for config in configs:
                    try:
                        text = pytesseract.image_to_string(processed, lang='jpn+eng', config=config).strip()
                        
                        if text:
                            # 获取置信度
                            try:
                                data = pytesseract.image_to_data(processed, lang='jpn+eng', config=config, output_type=pytesseract.Output.DICT)
                                confidences = [int(conf) for conf in data['conf'] if int(conf) > 0]
                                avg_confidence = sum(confidences) / len(confidences) if confidences else 0
                            except:
                                avg_confidence = 0
                            
                            logger.info(f"方法{i+1} OCR结果: '{text}' (置信度: {avg_confidence:.1f})")
                            
                            # 解析月份
                            month = self.parse_month_from_text(text)
                            if month and avg_confidence > best_confidence:
                                best_result = month
                                best_confidence = avg_confidence
                                logger.info(f"  -> 更新最佳结果: {month} (置信度: {avg_confidence:.1f})")
                    
                    except Exception as ocr_e:
                        continue
                        
            except Exception as method_e:
                continue
        
        if best_result:
            logger.info(f"月份识别结果: {best_result} (置信度: {best_confidence:.1f})")
        else:
            logger.warning(f"未能识别月份，位置: ({x},{y})")
            
        return best_result
    
    def parse_month_from_text(self, text: str) -> Optional[str]:
        """从文本中解析月份"""
        text = text.strip().replace(' ', '').replace('\n', '')
        
        # 直接匹配月份
        if text.endswith('月'):
            month_num = text[:-1]
            try:
                num = int(month_num)
                if 1 <= num <= 12:
                    return f"{num}月"
            except:
                pass
        
        # 匹配纯数字
        if text.isdigit():
            num = int(text)
            if 1 <= num <= 12:
                return f"{num}月"
        
        # 特殊处理10月（可能被识别为1O等）
        if text in ['10', '1O', 'IO', 'lo', 'lO']:
            return '10月'
        
        return None
    
    def process_pdf(self, pdf_path: str) -> List[Dict]:
        """处理PDF文件"""
        logger.info(f"开始处理PDF: {pdf_path}")
        
        doc = fitz.open(pdf_path)
        all_results = []
        
        for page_num in range(len(doc)):
            logger.info(f"处理第 {page_num + 1}/{len(doc)} 页")
            
            page = doc.load_page(page_num)
            
            # 高分辨率渲染
            mat = fitz.Matrix(2.0, 2.0)
            pix = page.get_pixmap(matrix=mat)
            img_data = pix.tobytes("ppm")
            
            nparr = np.frombuffer(img_data, np.uint8)
            image = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
            
            if image is None:
                logger.warning(f"第 {page_num + 1} 页图像转换失败")
                continue
            
            # 保存原始页面
            self.save_debug_image(image, f"page_{page_num + 1}_original.png")
            
            # 检测红色标记
            markings = self.detect_red_markings_improved(image)
            
            if not markings:
                logger.warning(f"第 {page_num + 1} 页未检测到有效标记")
                continue
            
            # 提取表格内容
            table_items = self.extract_table_content_improved(image, markings)
            
            # 提取月份信息并关联
            for item in table_items:
                mark = item['mark_position']
                month = self.extract_month_from_marking(image, mark)
                
                if month:
                    all_results.append({
                        'page': page_num + 1,
                        'note_number': item['note_number'],
                        'item_number': item['item_number'],
                        'month': month,
                        'mark_position': mark
                    })
                    logger.info(f"✅ 第{page_num + 1}页成功: {item['note_number']} {item['item_number']} -> {month}")
                else:
                    logger.warning(f"第{page_num + 1}页 {item['note_number']} {item['item_number']} 未识别到月份")
        
        doc.close()
        logger.info(f"PDF处理完成，共识别到 {len(all_results)} 个有效结果")
        return all_results
    
    def update_excel(self, excel_path: str, sheet_name: str, results: List[Dict]) -> int:
        """更新Excel文件"""
        logger.info(f"开始更新Excel文件: {excel_path}")
        
        try:
            wb = load_workbook(excel_path)
            
            if sheet_name not in wb.sheetnames:
                logger.error(f"工作表 '{sheet_name}' 不存在")
                return 0
            
            ws = wb[sheet_name]
            updated_count = 0
            
            for result in results:
                month = result['month']
                month_num = month.replace('月', '').zfill(2)
                date_str = f"2025-{month_num}"
                
                note_number = result['note_number']
                item_number = result['item_number']
                search_text = f"{note_number} {item_number}"
                
                # 查找对应行
                row_num = self.find_excel_row(ws, search_text)
                if row_num:
                    ws[f'O{row_num}'] = date_str
                    updated_count += 1
                    logger.info(f"更新 {search_text} -> {date_str}")
                else:
                    logger.warning(f"Excel中未找到: {search_text}")
            
            wb.save(excel_path)
            logger.info(f"Excel更新完成，共更新 {updated_count} 行")
            return updated_count
            
        except Exception as e:
            logger.error(f"Excel更新失败: {e}")
            return 0
    
    def find_excel_row(self, worksheet, search_text: str) -> Optional[int]:
        """在Excel中查找对应行"""
        for row in range(1, worksheet.max_row + 1):
            cell_value = str(worksheet[f'C{row}'].value or '').strip()
            if cell_value == search_text:
                return row
        return None


class GUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("PDF月份标记识别系统 - 精度改进版")
        self.root.geometry("700x600")
        
        self.recognizer = MonthMarkingRecognizer()
        self.pdf_path = None
        self.excel_path = None
        
        self.setup_ui()
    
    def setup_ui(self):
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Tesseract状态
        status_frame = ttk.Frame(main_frame)
        status_frame.grid(row=0, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
        
        try:
            version = pytesseract.get_tesseract_version()
            status_text = f"✅ Tesseract OCR {version} 已就绪"
            status_color = "green"
        except:
            status_text = "❌ Tesseract OCR 未配置"
            status_color = "red"
        
        ttk.Label(status_frame, text=status_text, foreground=status_color).pack()
        
        # 改进说明
        info_frame = ttk.Frame(main_frame)
        info_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
        ttk.Label(info_frame, text="🎯 精度改进版：优化了红色标记检测和表格内容提取", foreground="blue").pack()
        
        # 文件选择
        ttk.Label(main_frame, text="选择PDF文件:").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.pdf_label = ttk.Label(main_frame, text="未选择文件", foreground="gray")
        self.pdf_label.grid(row=2, column=1, sticky=tk.W, padx=10)
        ttk.Button(main_frame, text="浏览", command=self.select_pdf).grid(row=2, column=2, padx=5)
        
        ttk.Label(main_frame, text="选择Excel文件(可选):").grid(row=3, column=0, sticky=tk.W, pady=5)
        self.excel_label = ttk.Label(main_frame, text="未选择文件", foreground="gray")
        self.excel_label.grid(row=3, column=1, sticky=tk.W, padx=10)
        ttk.Button(main_frame, text="浏览", command=self.select_excel).grid(row=3, column=2, padx=5)
        
        ttk.Label(main_frame, text="选择工作表:").grid(row=4, column=0, sticky=tk.W, pady=5)
        self.sheet_var = tk.StringVar()
        self.sheet_combo = ttk.Combobox(main_frame, textvariable=self.sheet_var, state="readonly")
        self.sheet_combo.grid(row=4, column=1, sticky=(tk.W, tk.E), padx=10)
        
        # 调试选项
        debug_frame = ttk.Frame(main_frame)
        debug_frame.grid(row=5, column=0, columnspan=3, sticky=(tk.W))